import openpyxl
from selenium import webdriver; from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time, re

wb = openpyxl.load_workbook('Excel-Sheet.xlsx')
ws = wb['Sheet1']

# These print functions are for validating that you have selenium installed
# They return the spreadsheets information in your terminal
# def find_spreadsheet_range(ws):
#  print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
#  print('The value in cell A1 is: '+ws['A1'].value)
# find_spreadsheet_range(was)

base_url = "https://shop.COMPANY_WEBSITE.com/product/"

def remove_html_tags(get_price):
    clean = re.compile('<.*?>')
    return re.sub(clean, '', get_price)

def write_price(value):
    f = open('updated_price.html', 'a')
    product_url = base_url+str(value)
  # eldir = Element Directory; You can find out the element directory by inspecting the page you wish to scrape
  # and then copy the HTML Path to the line of HTML that you want to extract. 
  # In this example, the Price Element can be found at eldir.
    eldir = '//*[@id="product-details"]/div[1]/section/div/div/div/div[1]/ul[2]/li[2]/span[1]'
    options = Options()
    # Enable Headless Mode
    options.add_argument = ('--headless=new')

    driver = webdriver.Chrome(options=options, service=ChromeService(ChromeDriverManager().install()))
    driver.get(product_url)
    time.sleep(5)
    get_price = driver.find_element(By.XPATH, eldir).get_attribute('outerHTML')
    time.sleep(5)
    driver.close()
    print(remove_html_tags(get_price))
    f.write(remove_html_tags(get_price)+'\n')
    f.close()
    time.sleep(10)

def write_blank_price():
    print("SKU is not in COMPANY. Please try again.")
    f = open('updated_price.html', 'a')
    f.write('$0.00\n')
    f.close()

# When ready to execute code, replace all "your_num" values with the range of cells you wish to search in your spreadsheet.
# If you do not know the range of cells, run the find_spreadsheet_ranges function independently and then add in the values necessary.
# Ideally, min_row would be 1, max_column would be 1 & max_row would be the last line in your spreadsheet.
def value_gathering():
    for row in ws.iter_cols(min_row=your_num, max_col=your_num, max_row=your_num, values_only=True):
        for value in row:
            try:
                if str(value) == "NULL":
                    print(str(value)+" The SKU read is 'NULL', please try again.")
                    write_blank_price()
                elif str(value) == "None":
                    print(str(value)+" The SKU read is 'NONE', please try again.")
                    write_blank_price()
                else:
                    print(str(value)+": Thank you for inputting the correct SKU. Give us a moment to find the price and add it to your Spreadsheet.")

                    write_price(value)
                    time.sleep(5)

            except Exception as oopsies:
                print("An error has occured, please give us a moment.")
                print(value, ": Cannot be read. Error Code:", oopsies )
                f = open('updated_price.html', 'a')
                f.write('$0.00\n')
                f.close()
                time.sleep(2)
                continue
value_gathering()
